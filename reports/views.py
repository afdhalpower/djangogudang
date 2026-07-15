"""
Reports — Inventory, Low Stock, Stock Card, and CSV export.
"""
import csv
from io import StringIO
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.db.models import F, Sum, Q, Count, ExpressionWrapper, DecimalField
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views.generic import TemplateView, ListView, DetailView
from products.models import Product
from stock.models import StockTransaction, StockTransactionItem
from suppliers.models import Supplier

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Helpers ───────────────────────────────────────────────────

def _csv_response(filename: str, header: list, rows: list) -> HttpResponse:
    """Return an HttpResponse with CSV content (Content-Type + download headers)."""
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    return response


def _xlsx_response(filename: str, header: list, rows: list) -> HttpResponse:
    """Return an HttpResponse with formatted Excel sheet content using emerald theme."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    ws.append(header)
    for col_num in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align
        cell.border = thin_border
    
    for row_num, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            
            val = cell.value
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = right_align
            elif str(val).startswith("Rp") or str(val).isdigit():
                cell.alignment = right_align
            else:
                cell.alignment = left_align
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def _format_rp(val):
    """Format number to IDR string without symbol (for CSV numbers)."""
    if val is None:
        return 0
    return int(val)


# ─── Inventory Report ──────────────────────────────────────────

class InventoryReportView(LoginRequiredMixin, TemplateView):
    """
    Full product inventory table: current stock, minimum, status,
    purchase/selling price, and total inventory value.
    """
    template_name = "reports/inventory.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        products = Product.objects.select_related("category", "unit", "supplier").annotate(
            inventory_value=ExpressionWrapper(
                F("current_stock") * F("purchase_price"),
                output_field=DecimalField()
            )
        )

        # Totals calculated at database level
        totals = Product.objects.aggregate(
            total_inventory=Sum(
                ExpressionWrapper(F("current_stock") * F("purchase_price"), output_field=DecimalField())
            ),
            total_selling=Sum(
                ExpressionWrapper(F("current_stock") * F("selling_price"), output_field=DecimalField())
            ),
        )
        total_inventory_value = totals["total_inventory"] or 0
        total_selling_value = totals["total_selling"] or 0

        total_items = products.count()
        low_stock_count = products.filter(current_stock__lte=F("minimum_stock")).count()

        context.update(
            products=products,
            total_inventory_value=total_inventory_value,
            total_selling_value=total_selling_value,
            total_items=total_items,
            low_stock_count=low_stock_count,
        )
        return context


@login_required
def inventory_csv(request):
    """Export Inventory Report as CSV."""
    products = Product.objects.select_related("category", "unit", "supplier").all()
    rows = []
    for p in products:
        rows.append([
            p.sku, p.name, p.category.name if p.category else "",
            p.unit.abbreviation if p.unit else "",
            p.current_stock, p.minimum_stock,
            p.supplier.company_name if p.supplier else "",
            _format_rp(p.purchase_price),
            _format_rp(p.selling_price),
            _format_rp(p.current_stock * p.purchase_price),
            p.status,
        ])
    return _csv_response(
        "inventory_report",
        ["SKU", "Name", "Category", "Unit", "Stock", "Min", "Supplier",
         "Purchase Price", "Selling Price", "Inventory Value", "Status"],
        rows,
    )


@login_required
def inventory_xlsx(request):
    """Export Inventory Report as Excel."""
    products = Product.objects.select_related("category", "unit", "supplier").all()
    rows = []
    for p in products:
        rows.append([
            p.sku, p.name, p.category.name if p.category else "",
            p.unit.abbreviation if p.unit else "",
            p.current_stock, p.minimum_stock,
            p.supplier.company_name if p.supplier else "",
            _format_rp(p.purchase_price),
            _format_rp(p.selling_price),
            _format_rp(p.current_stock * p.purchase_price),
            p.status,
        ])
    return _xlsx_response(
        "inventory_report",
        ["SKU", "Name", "Category", "Unit", "Stock", "Min", "Supplier",
         "Purchase Price", "Selling Price", "Inventory Value", "Status"],
        rows,
    )


# ─── Low Stock Report ──────────────────────────────────────────

class LowStockReportView(LoginRequiredMixin, TemplateView):
    """All products where current_stock <= minimum_stock."""
    template_name = "reports/low_stock.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = Product.objects.filter(
            current_stock__lte=F("minimum_stock")
        ).select_related("category", "unit", "supplier").order_by("current_stock")
        context["products"] = products
        context["count"] = products.count()

        # Compile reorder recommendations grouped by supplier
        from collections import defaultdict
        recommendations = defaultdict(list)
        for p in products:
            if p.supplier:
                shortage = p.minimum_stock - p.current_stock
                # Default recommended order quantity: shortage doubled, min 10 units
                reorder_qty = max(shortage * 2, 10)
                recommendations[p.supplier].append({
                    "product": p,
                    "shortage": shortage,
                    "reorder_qty": reorder_qty,
                })

        reorder_list = []
        for supplier, items in recommendations.items():
            email_body = f"Dear {supplier.company_name},\n\nWe would like to request a reorder for the following items:\n\n"
            for item in items:
                prod = item["product"]
                email_body += f"- {prod.name} (SKU: {prod.sku}) — Qty: {item['reorder_qty']} {prod.unit.abbreviation}\n"
            email_body += "\nPlease confirm the pricing and estimated delivery date.\n\nBest regards,\nWarehouse Management"

            reorder_list.append({
                "supplier": supplier,
                "items": items,
                "email_subject": f"Purchase Reorder - {supplier.company_name}",
                "email_body": email_body,
            })

        context["reorder_recommendations"] = reorder_list
        return context


@login_required
def low_stock_csv(request):
    """Export Low Stock Report as CSV."""
    products = Product.objects.filter(
        current_stock__lte=F("minimum_stock")
    ).select_related("category", "unit", "supplier").order_by("current_stock")
    rows = []
    for p in products:
        rows.append([
            p.sku, p.name, p.category.name if p.category else "",
            p.unit.abbreviation if p.unit else "",
            p.current_stock, p.minimum_stock,
            p.supplier.company_name if p.supplier else "",
            _format_rp(p.purchase_price),
            _format_rp(p.selling_price),
            p.status,
        ])
    return _csv_response(
        "low_stock_report",
        ["SKU", "Name", "Category", "Unit", "Stock", "Min", "Supplier",
         "Purchase Price", "Selling Price", "Status"],
        rows,
    )


@login_required
def low_stock_xlsx(request):
    """Export Low Stock Report as Excel."""
    products = Product.objects.filter(
        current_stock__lte=F("minimum_stock")
    ).select_related("category", "unit", "supplier").order_by("current_stock")
    rows = []
    for p in products:
        rows.append([
            p.sku, p.name, p.category.name if p.category else "",
            p.unit.abbreviation if p.unit else "",
            p.current_stock, p.minimum_stock,
            p.supplier.company_name if p.supplier else "",
            _format_rp(p.purchase_price),
            _format_rp(p.selling_price),
            p.status,
        ])
    return _xlsx_response(
        "low_stock_report",
        ["SKU", "Name", "Category", "Unit", "Stock", "Min", "Supplier",
         "Purchase Price", "Selling Price", "Status"],
        rows,
    )


# ─── Stock Card ────────────────────────────────────────────────

class StockCardView(LoginRequiredMixin, TemplateView):
    """
    Per-product movement history — shows every IN/OUT/ADJUSTMENT
    that affected this product's stock.
    """
    template_name = "reports/stock_card.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product_id = self.request.GET.get("product", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")

        products = Product.objects.filter(status="active").select_related("unit")
        context["products"] = products

        if product_id and product_id.isdigit():
            product = get_object_or_404(Product, pk=int(product_id))
            context["selected_product"] = product

            items = StockTransactionItem.objects.filter(
                product=product
            ).select_related("transaction", "product__unit").order_by(
                "-transaction__date", "-transaction__created_at"
            )

            if date_from:
                items = items.filter(transaction__date__gte=date_from)
            if date_to:
                items = items.filter(transaction__date__lte=date_to)

            context["items"] = items
            context["date_from"] = date_from
            context["date_to"] = date_to

        return context


@login_required
def stock_card_csv(request):
    """Export Stock Card for a specific product as CSV."""
    product_id = request.GET.get("product", "")
    if not product_id or not product_id.isdigit():
        return HttpResponse("Select a valid product first", status=400)

    product = get_object_or_404(Product, pk=int(product_id))
    items = StockTransactionItem.objects.filter(
        product=product
    ).select_related("transaction").order_by("transaction__date")

    rows = []
    for item in items:
        movement = item.transaction.movement_type
        direction = item.transaction.adjustment_direction if movement == "adjustment" else ""
        rows.append([
            item.transaction.date,
            item.transaction.reference_number or "",
            movement.upper(),
            direction,
            item.quantity,
            _format_rp(item.unit_price),
        ])

    return _csv_response(
        f"stock_card_{product.sku}",
        ["Date", "Reference", "Type", "Direction", "Qty", "Unit Price"],
        rows,
    )


@login_required
def stock_card_xlsx(request):
    """Export Stock Card for a specific product as Excel."""
    product_id = request.GET.get("product", "")
    if not product_id or not product_id.isdigit():
        return HttpResponse("Select a valid product first", status=400)

    product = get_object_or_404(Product, pk=int(product_id))
    items = StockTransactionItem.objects.filter(
        product=product
    ).select_related("transaction").order_by("transaction__date")

    rows = []
    for item in items:
        movement = item.transaction.movement_type
        direction = item.transaction.adjustment_direction if movement == "adjustment" else ""
        rows.append([
            item.transaction.date.strftime("%Y-%m-%d") if item.transaction.date else "",
            item.transaction.reference_number or "",
            movement.upper(),
            direction,
            item.quantity,
            _format_rp(item.unit_price),
        ])

    return _xlsx_response(
        f"stock_card_{product.sku}",
        ["Date", "Reference", "Type", "Direction", "Qty", "Unit Price"],
        rows,
    )
